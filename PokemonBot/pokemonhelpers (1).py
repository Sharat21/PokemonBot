'''
Ordering of pokemon data in xlsx file

1) Name
2) Stats
3) Ivy
4) Nature
'''

def CreateFile(guildID: str) -> None:
    '''
    Create a file for the guild (initialization)
    '''
    from openpyxl import Workbook
    workbook = Workbook()
    workbook.active.title = "general"
    workbook.save(guildID + '.xlsx')

def CatchPokemon(pokemonInfo: str, guildID: str, playerID: str) -> None:
    '''
    Capture a pokemon and save it to a players inventory
    '''
    from openpyxl import load_workbook
    import pokebase as pb

    pokemonInfo = stats(pb.pokemon(pokemonInfo.lower().strip()))
    wb = load_workbook(guildID + ".xlsx")
    wb.active = wb[playerID]
    wb.active.append(pokemonInfo)
    wb.save(guildID + ".xlsx")

def stats(pokemon) -> list[str]:
    '''
    Get slightly randomized stats of a specific pokemon
    '''
    from random import randint, choice
    from pokemonconstants import NATURES
    items = []
    items.append(pokemon.name.capitalize())
    items.append(str([x.stat.name + ":" + str(randint(x.base_stat - 5, x.base_stat + 5)) for x in pokemon.stats]))
    items.append(str([x.stat.name + ":" + str(randint(0, 31)) for x in pokemon.stats]))
    items.append(choice(NATURES))
    return items

def get_player_pokemon(guildID: str, playerID: str, index: int) -> list: #Pass in 0 for index if you are just starting a fight
    '''
    Get information about a pokemon that a player owns
    '''
    from openpyxl import load_workbook
    sheet = load_workbook(guildID + ".xlsx")[playerID]
    if index == 0:
        from openpyxl import load_workbook
        index = int(sheet['B1'].value)
    index += 1
    return [sheet["A" + str(index)].value, sheet["B" + str(index)].value, sheet["C" + str(index)].value,sheet["D" + str(index)].value]

async def fight(pokemon, opponent, turn, ctx, client):
    '''
    Begin a pokemon battle
    '''
    from random import randint
    def check(m): #Ignore this
        return m.author == ctx.author and m.channel.id == ctx.channel.id

    if pokemon.is_dead():
        await ctx.send("Your pokemon fainted.")
        return
    elif opponent.is_dead():
        AddWin(str(ctx.guild.id), str(ctx.author.id))
        await ctx.send("Opponent has fainted.")
        if catch(opponent.health):
            await ctx.send("Pokemon has been caught.")
            CatchPokemon(opponent.name, str(ctx.guild.id), str(ctx.author.id))
        else:
            await ctx.send("Pokemon has escaped.")
        return
    else:
        await ctx.send(f"{pokemon.name}'s health: {pokemon.health} | {opponent.name}'s health: {opponent.health} ")
        if turn:
            await ctx.send("Players turn")
            await ctx.send("What would you like to do? (flee, catch, attack, special)")
            input = await client.wait_for('message', timeout = 60, check=check)
            input = input.content.strip().lower()
            if input == 'flee':
                return

            elif input == 'catch':
                if catch(opponent.health):
                    AddWin(str(ctx.guild.id), str(ctx.author.id))
                    await ctx.send("Pokemnon has been caught.")
                    CatchPokemon(opponent.name, str(ctx.guild.id), str(ctx.author.id))
                    return
                else:
                    await ctx.send("Pokemon was not caught.")
                    chance = randint(0,1)
                    if chance == 0:
                        await fight(pokemon, opponent, not turn, ctx, client)
                    else:
                        await ctx.send("Pokemon has escaped.")
                        return

            else:
                attack = pokemon.deal_damage(input)
                if not attack:
                    await ctx.send(pokemon.name + " tried using their special move, but failed.")
                else:
                    opponent.take_damage(attack)

                await pokemon.special_cooldown(ctx)
                await fight(pokemon, opponent, not turn, ctx, client)

        else:
            await ctx.send("Opponents turn")
            if opponent.have_special:
                pokemon.take_damage(opponent.deal_damage('special'))

            else:
                pokemon.take_damage(opponent.deal_damage(" "))
            await opponent.special_cooldown(ctx)
            await fight(pokemon, opponent, not turn, ctx, client)

def AddWin(guildiD, playerid) -> None:
    '''
    Update the amount of wins a player has against pokemon in the wild
    '''
    from openpyxl import load_workbook
    wb = load_workbook(guildiD + ".xlsx")
    ws = wb[playerid]
    wins = int(ws['D1'].value)
    ws['D1'] = str(wins + 1)
    wb.save(guildiD + "xlsx")


async def hunting(habitat : str, player, ctx, client) -> None:
    '''
    Start a hunting event to find a new pokemon
    '''
    import pokebase as pb
    from random import choice, randint
    from pokemonobject import Pokemon

    opponent = choice(pb.pokemon_habitat(habitat).pokemon_species).name
    opponent = pb.pokemon(opponent)
    opponent = Pokemon(
        opponent.name,
        opponent.stats[0].base_stat,
        opponent.stats[1].base_stat,
        opponent.stats[2].base_stat,
        opponent.stats[3].base_stat,
    )


    playerPokemon = Pokemon(
        player[0],
        int(player[1][player[1].find('hp') + 3:player[1].find(',') - 1]),
        int(player[1][player[1].find('attack') + 7:player[1].find(',', player[1].find('attack')) - 1]),
        int(player[1][player[1].find('defense') + 8:player[1].find(',', player[1].find('defense')) - 1]),
        int(player[1][player[1].find('special-attack') + 15:player[1].find(',', player[1].find('special-attack')) - 1]),
    )
    await fight(playerPokemon, opponent, bool(randint(0, 1)), ctx, client)

def catch(health: int):
    '''
    Attempt to catch a pokemon
    '''
    from random import randint
    chance = randint(60, 80) #Max chance is second number
    return chance - health >= randint(0, 100)