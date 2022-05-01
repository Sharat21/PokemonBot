import discord
from discord.ext import commands

client=commands.Bot(command_prefix = '!', case_insensitive=True)

@client.event
async def on_ready():
    '''
    Bot has started running -> Debug tool
    '''
    print('We have logged in as {0.user}'.format(client))


@client.event
async def on_guild_join(guild):
    '''
    Create xlsx file for pokemon information when joining a new server
    '''
    from pokemonhelpers import CreateFile
    CreateFile(str(guild.id))

@client.command(aliases = ['info'])
async def pokemoninfo(ctx, pokemoninput : str):
    '''
    Get information about a specific pokemon
    '''
    import pokebase as pb
    pokemon = pb.pokemon(pokemoninput.lower())
    species = pb.pokemon_species(pokemon.name)

    if species.is_legendary or species.is_mythical:
        color = discord.Color.gold()

    else:
        color = discord.Color.default()

    if species.evolves_from_species is not None:
        evolution = "Evolves from: " + species.evolves_from_species.name.capitalize() + "\n"

    else:
        evolution = ""

    embed = discord.Embed(title = pokemon.name.capitalize() + "  #" + str(pokemon.id).zfill(3), description = (
        species.flavor_text_entries[6].flavor_text.replace("\n", " ") + "\n\n"
        "Types: " + ", ".join([x.type.name.capitalize().replace("-", " ") for x in pokemon.types]) + '\n' +
        "Abilities: " + ", ".join([x.ability.name.capitalize().replace("-", " ") for x in pokemon.abilities]) + "\n" +
        "Egg Groups: " + ", ".join([x.name.capitalize() for x in species.egg_groups]) + "\n" +
        evolution +
        "Base Experience: " + str(pokemon.base_experience) + "\n" +
        "Height: " + str(pokemon.height / 10) + "\n" +
        "Weight: " + str(pokemon.weight)
    ), color= color)
    embed.set_thumbnail(url=pokemon.sprites.front_default)

    for stat in pokemon.stats:
        embed.add_field(name= stat.stat.name.replace("-", " "), value = stat.base_stat, inline=True)
    await ctx.send(embed = embed)

@client.command(aliases = ['moves'])
async def pokemonmoves(ctx, pokemoninput : str):
    '''
    Get a message containing all the moves a pokemon can use
    '''
    import pokebase as pb
    pokemon = pb.pokemon(pokemoninput.lower())
    embed = discord.Embed(title = pokemon.name.capitalize() + "  #" + str(pokemon.id).zfill(3), description = "This pokemon can learn the following abilities")
    embed.set_thumbnail(url=pokemon.sprites.front_default)

    for move in pokemon.moves:
        embed.add_field(name= "\u200b", value = move.move.name.replace("-", " ") , inline=True)
    await ctx.send(embed = embed)

@client.command(aliases =['move'])
async def moveinfo(ctx, *moveinput: str):
    '''
    Get information about a specific move'''
    import pokebase as pb
    moveS = ""
    for word in moveinput:
        moveS += word + "-"

    move = pb.move(moveS[0:-1].lower())

    embed = discord.Embed(title = move.name.replace("-", " ").capitalize(), description = (
        move.flavor_text_entries[0].flavor_text.replace("\n", " ") + "\n\n" +
        move.effect_entries[0].effect.replace("\n", " ") + "\n" +
        "Type: " + move.type.name.capitalize() + "\n" +
        "Accuracy: " + str(move.accuracy) + "\n" +
        "Power: " + str(move.power) + "\n" +
        "PP: " + str(move.pp) + "\n" +
        "Priority: " + str(move.priority) + "\n"
        "Effect Chance: " + str(move.effect_chance)
    ))
    await ctx.send(embed = embed)

@client.command(aliases = ['begin'])
async def start(ctx, pokemon):
    '''
    Initialize the pokemon game for a player
    '''
    from os.path import exists
    from pokemonhelpers import CatchPokemon

    if pokemon == "":
        await ctx.send ("You must choose a starter pokemon")
        return

    if not exists(str(ctx.guild.id) + ".xlsx"):
        from pokemonhelpers import CreateFile
        CreateFile(str(ctx.guild.id))

    from openpyxl import load_workbook
    wb = load_workbook(str(ctx.guild.id) + ".xlsx")

    if str(ctx.author.id) in wb.sheetnames:
        await ctx.send("You already picked a starter pokemon")
        return

    wb.create_sheet(str(ctx.author.id))
    ws = wb[str(ctx.author.id)]
    ws.append(["Default: ", 1, "Wins: ", 0])
    wb.save(str(ctx.guild.id) + ".xlsx")

    if pokemon.lower().strip() == "squirtle":
        CatchPokemon("squirtle", str(ctx.guild.id), str(ctx.author.id))

    elif pokemon.lower().strip() == "charmander":
        CatchPokemon("charmander", str(ctx.guild.id), str(ctx.author.id))

    else:
        CatchPokemon('bulbasaur', str(ctx.guild.id), str(ctx.author.id))
    await ctx.send("Pokemon chosen! Good luck adventurer.")

@client.command(aliases = ['adventure', 'hunting'])
async def hunt(ctx, *region):
    '''
    Start hunting for a pokemon
    '''
    from pokemonhelpers import hunting, get_player_pokemon
    from pokemonconstants import HABITATS
    if region == () or region[0].strip().lower() not in HABITATS:
        from random import choice
        region = (choice(HABITATS),)
    player = get_player_pokemon(str(ctx.guild.id), str(ctx.author.id), 0)
    await hunting(region[0].strip().lower(), player,ctx, client)

@client.command(aliases = ['my-pokemon', 'my pokemom', 'pokemon'])
async def mypokemon(ctx, *pokemon: str):
    '''
    Get a list of pokemon you own, or specific information about one if its index is provided
    '''
    from openpyxl import load_workbook
    wb = load_workbook(str(ctx.guild.id) + ".xlsx")
    ws = wb[str(ctx.author.id)]

    if len(pokemon) == 0 or not pokemon[0].isdigit():
        acc = ""
        for index, pokemon in enumerate(ws['A']):
            acc += str(index) + "  |  " + str(pokemon.value) + "\n"

        acc = acc[acc.find("\n"):]
        embed = discord.Embed(title = "My Pokemon", description = acc)
        await ctx.send(embed = embed)

    else:
        try:
            from pokemonhelpers import get_player_pokemon
            player = get_player_pokemon(str(ctx.guild.id), str(ctx.author.id), int(pokemon[0]))
            embed = discord.Embed(title = player[0].replace("-", " ").capitalize(), description = (
                "Base Stats \n" +
                "HP: " + player[1][player[1].find('hp') + 3:player[1].find(',') - 1] + "\n"+
                "Attack: " + player[1][player[1].find('attack') + 7:player[1].find(',', player[1].find('attack')) - 1] + "\n" +
                "Defense: " + player[1][player[1].find('defense') + 8:player[1].find(',', player[1].find('defense')) - 1]+ "\n" +
                "Special Attack: " + player[1][player[1].find('special-attack') + 15:player[1].find(',', player[1].find('special-attack')) - 1]+ "\n"
                "Special Defense: " + player[1][player[1].find('special-defense') + 16:player[1].find(',', player[1].find('special-defense')) - 1]+ "\n" +
                "Speed: " + player[1][player[1].find('special-attack') + 15:player[1].find(',', player[1].find('speed')) - 1]+ "\n" +
                "\n Ivy Stats \n" +
                "HP: " + player[2][player[2].find('hp') + 3:player[2].find(',') - 1] + "\n"+
                "Attack: " + player[2][player[2].find('attack') + 7:player[2].find(',', player[2].find('attack')) - 2] + "\n" +
                "Defense: " + player[2][player[2].find('defense') + 8:player[2].find(',', player[2].find('defense')) - 2]+ "\n" +
                "Special Attack: " + player[2][player[2].find('special-attack') + 15:player[2].find(',', player[1].find('special-attack')) - 2]+ "\n"
                "Special Defense: " + player[2][player[2].find('special-defense') + 16:player[2].find(',', player[1].find('special-defense')) - 2]+ "\n" +
                "Speed: " + player[2][player[2].find('special-attack') + 15:player[2].find(',', player[2].find('speed')) - 2]+ "\n\n" +
                "Nature: " + player[3]
            ))
            await ctx.send(embed = embed)

        except:
            await ctx.send("Unable to find pokemon at this index")

@client.command(aliases = ['victories'])
async def wins(ctx):
    from openpyxl import load_workbook

    ws = load_workbook(str(ctx.guild.id) + ".xlsx")[str(ctx.author.id)]

    await ctx.send("Wins: " + str(ws['D1'].value))

client.run('OTY5Nzk1OTM0NzMwOTc3MzIx.Ymym2A.1nbeLvUomdkpA20zKYHL7sJahZU')